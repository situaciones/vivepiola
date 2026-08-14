import io

import openpyxl
import openpyxl.comments
import openpyxl.utils
import openpyxl.worksheet.datavalidation
from django.core.exceptions import ValidationError
from django.core.validators import validate_email

from .models import (
    CondicionEspecial, Permanencia, Persona, RolOcupacion, Unidad, VinculoCopropietario,
)

COLUMNAS_PLANTILLA = [
    'unidad', 'rol_ocupacion', 'nombre_completo', 'cedula_identidad',
    'domicilio', 'correo_electronico', 'telefono', 'permanencia', 'vinculo_copropietario',
    'condicion_especial',
]

ROLES_VALIDOS = {choice.value for choice in RolOcupacion}
PERMANENCIAS_VALIDAS = {choice.value for choice in Permanencia}
VINCULOS_VALIDOS = {choice.value for choice in VinculoCopropietario}
CONDICIONES_VALIDAS = {choice.value for choice in CondicionEspecial}


# Ayuda por columna: se muestra al pararse en la celda, sin abrir otra hoja.
AYUDA_COLUMNAS = {
    'unidad': ('Obligatorio', 'El departamento, casa o estacionamiento. Ej: Depto 302'),
    'rol_ocupacion': ('Obligatorio', 'Titulo con que ocupa la unidad. Elija de la lista.'),
    'nombre_completo': ('Obligatorio', 'Nombre y apellidos de la persona.'),
    'cedula_identidad': ('Obligatorio', 'RUT con guion. Ej: 12.345.678-9'),
    'domicilio': ('Obligatorio', 'Direccion de la unidad.'),
    'correo_electronico': (
        'Obligatorio',
        'Es el canal legal de notificacion: sin correo valido la multa no se puede notificar.',
    ),
    'telefono': ('Opcional', 'Formato +569XXXXXXXX. Habilita el aviso por WhatsApp.'),
    'permanencia': (
        'Opcional',
        'PERMANENTE si no se indica. TRANSITORIO (estadia corta) hace que la '
        'notificacion se copie tambien al propietario.',
    ),
    'vinculo_copropietario': (
        'Opcional',
        'Solo si ocupa por su vinculo con el dueño. En ese caso la notificacion '
        'tambien se le copia al propietario.',
    ),
    'condicion_especial': (
        'Opcional',
        'Solo si corresponde: FALLECIDO, DISCAPACIDAD o REQUIERE_APOYO. Detiene el '
        'cobro automatico para que el Comite revise el caso antes de cobrarlo.',
    ),
}

# Columnas con valores cerrados: van como lista desplegable para que no haya
# que adivinar la ortografia ni acordarse de escribir en mayusculas.
LISTAS_DESPLEGABLES = {
    'condicion_especial': sorted(c for c in CONDICIONES_VALIDAS if c),
    'rol_ocupacion': sorted(ROLES_VALIDOS),
    'permanencia': sorted(PERMANENCIAS_VALIDAS),
    'vinculo_copropietario': sorted(v for v in VINCULOS_VALIDOS if v),
}

OBLIGATORIAS = {c for c, (marca, _) in AYUDA_COLUMNAS.items() if marca == 'Obligatorio'}


def generar_plantilla_excel():
    """
    Plantilla .xlsx para el administrador.

    Se abre en la hoja de instrucciones a proposito: quien la descarga por
    primera vez necesita leer como llenarla antes de ver una grilla vacia.
    Ademas cada encabezado lleva su ayuda y las columnas de valores cerrados
    van como lista desplegable, para que la guia este donde se escribe.

    Va sin filas de ejemplo: una plantilla con datos de muestra termina
    importando personas ficticias al registro.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Registro Copropietarios'
    ws.append(COLUMNAS_PLANTILLA)
    ws.freeze_panes = 'A2'
    for i, ancho in enumerate([16, 16, 30, 18, 38, 30, 18, 16, 22, 22], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = ancho

    for i, columna in enumerate(COLUMNAS_PLANTILLA, start=1):
        letra = openpyxl.utils.get_column_letter(i)
        celda = ws.cell(row=1, column=i)
        celda.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        # Obligatorias en oscuro, opcionales en gris: se distingue de un vistazo
        # cuales no se pueden dejar en blanco.
        celda.fill = openpyxl.styles.PatternFill(
            'solid', fgColor='111827' if columna in OBLIGATORIAS else '6B7280',
        )
        marca, detalle = AYUDA_COLUMNAS[columna]
        celda.comment = openpyxl.comments.Comment(f'{marca}. {detalle}', 'VIVEPIOLA')

        rango = f'{letra}2:{letra}1000'
        if columna in LISTAS_DESPLEGABLES:
            opciones = LISTAS_DESPLEGABLES[columna]
            validacion = openpyxl.worksheet.datavalidation.DataValidation(
                type='list', formula1='"' + ','.join(opciones) + '"', allow_blank=True,
                showDropDown=False,  # False = SI muestra la flecha (bandera invertida en el formato)
                errorTitle='Valor no valido',
                error=f'Use uno de estos valores: {", ".join(opciones)}',
                promptTitle=columna, prompt=detalle,
            )
        else:
            validacion = openpyxl.worksheet.datavalidation.DataValidation(
                type=None, allow_blank=True, promptTitle=columna, prompt=f'{marca}. {detalle}',
            )
        validacion.showInputMessage = True
        ws.add_data_validation(validacion)
        validacion.add(rango)

    guia = wb.create_sheet('Instrucciones')
    # Los titulos se marcan con TITULO en vez de por numero de fila: la lista
    # de indices fijos que habia antes ponia negrita en la linea equivocada
    # apenas se agregaba un renglon en el medio.
    TITULO = object()
    for fila in [
        (TITULO, 'COMO LLENAR ESTA PLANILLA', ''),
        (None, '', ''),
        (None, 'Los datos van en la otra hoja, la de abajo: "Registro Copropietarios".', ''),
        (None, 'Una fila por PERSONA, no por departamento.', ''),
        (None, 'Si un depto tiene propietario y arrendatario, van dos filas con la misma unidad.', ''),
        (None, '', ''),
        (None, 'Encabezado oscuro = obligatorio. Encabezado gris = opcional.', ''),
        (None, 'Al pararse en una celda aparece la ayuda de esa columna.', ''),
        (None, 'rol_ocupacion, permanencia, vinculo y condicion son listas desplegables:', ''),
        (None, '   elija el valor y no tendra que acordarse de escribirlo en mayusculas.', ''),
        (None, '', ''),
        (TITULO, 'COLUMNA', 'QUE VA / REGLAS'),
        (None, 'unidad', 'Ej: Depto 302, Estacionamiento 12'),
        (None, 'rol_ocupacion', 'Titulo con que ocupa: ' + ', '.join(sorted(ROLES_VALIDOS))),
        (None, 'nombre_completo', 'Nombre y apellidos'),
        (None, 'cedula_identidad', 'RUT con guion. Ej: 12.345.678-9'),
        (None, 'domicilio', 'Direccion de la unidad'),
        (None, 'correo_electronico', 'OBLIGATORIO: es el canal legal de notificacion'),
        (None, 'telefono', 'Opcional. +569XXXXXXXX. Habilita el aviso por WhatsApp'),
        (None, 'permanencia',
         'Opcional. PERMANENTE (por defecto) o TRANSITORIO (hospedaje temporal, estadia corta). '
         'Marcar TRANSITORIO hace que la notificacion se copie al propietario.'),
        (None, 'vinculo_copropietario',
         'Opcional. CONYUGE, CONVIVIENTE_CIVIL o FAMILIAR. Si ocupa por vinculo con el dueño, '
         'la notificacion tambien se le copia a el.'),
        (None, 'condicion_especial',
         'Opcional. FALLECIDO, DISCAPACIDAD o REQUIERE_APOYO. Detiene el cobro automatico '
         'para que el Comite confirme antes de cobrar.'),
        (None, '', ''),
        (TITULO, 'QUIEN PAGA', ''),
        (None, 'El PROPIETARIO es el obligado al pago ante la comunidad.', ''),
        (None, 'Si multan al arrendatario, el cargo igual se emite a nombre del dueño de la unidad.', ''),
        (None, 'Por eso cada unidad deberia tener su propietario registrado.', ''),
        (None, '', ''),
        (TITULO, 'A QUIEN LE LLEGA LA NOTIFICACION', ''),
        (None, 'La notificacion legal va al correo del infractor. Ademas se copia al propietario cuando:', ''),
        (None, '   - la persona es TRANSITORIO: puede haberse ido antes de que venza el plazo de descargo.', ''),
        (None, '   - la persona ocupa por vinculo con el dueño (conyuge, conviviente civil, familiar).', ''),
        (None, 'En los demas casos no se copia a nadie mas.', ''),
        (None, 'Por eso conviene llenar estas dos columnas: son las que definen a quien le llega.', ''),
        (None, '', ''),
        (TITULO, 'ERRORES QUE RECHAZAN LA FILA', ''),
        (None, 'Correo vacio o mal escrito', 'Sin correo no se puede notificar y la multa queda bloqueada'),
        (None, 'rol_ocupacion fuera de la lista', 'En mayusculas y sin acentos'),
        (None, 'Falta nombre, cedula, domicilio o unidad', 'Todos obligatorios'),
        (None, '', ''),
        (None, 'El sistema valida fila por fila: importa las correctas e informa cuales fallaron.', ''),
        (None, 'Puedes volver a subir el archivo corregido: las personas ya cargadas se actualizan.', ''),
    ]:
        marca, columna_a, columna_b = fila
        guia.append([columna_a, columna_b])
        if marca is TITULO:
            for celda in guia[guia.max_row]:
                celda.font = openpyxl.styles.Font(bold=True)
    guia.column_dimensions['A'].width = 62
    guia.column_dimensions['B'].width = 62

    # Se abre en las instrucciones, no en la grilla vacia. El importador busca
    # la hoja del registro por sus encabezados (ver _hoja_de_datos), asi que
    # esto no afecta la carga aunque el administrador vuelva a guardar aqui.
    wb.move_sheet(guia, offset=-1)
    wb.active = 0
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _hoja_de_datos(wb):
    """
    Busca la hoja que trae el registro, no la que el usuario dejo seleccionada.

    Excel guarda como hoja activa la ultima que estuvo a la vista. Leer wb.active
    significaba que, si el administrador cerraba el archivo mirando las
    instrucciones, la importacion intentaba leer las instrucciones como si
    fueran personas. Se busca por encabezados y solo se cae en la activa si
    ninguna hoja los tiene (archivo armado a mano).
    """
    obligatorios = {'unidad', 'nombre_completo'}
    for hoja in wb.worksheets:
        primera = next(hoja.iter_rows(max_row=1, values_only=True), ())
        encabezados = {str(c).strip().lower() for c in primera if c}
        if obligatorios <= encabezados:
            return hoja
    return wb.active


def _leer_filas(archivo):
    """Soporta .xlsx y .csv; devuelve lista de dicts normalizados por encabezado."""
    nombre = archivo.name.lower()
    if nombre.endswith('.csv'):
        import csv
        texto = archivo.read().decode('utf-8-sig')
        lector = csv.DictReader(io.StringIO(texto))
        filas = [
            {(k or '').strip().lower(): (v or '').strip() for k, v in fila.items()}
            for fila in lector
        ]
        return filas

    wb = openpyxl.load_workbook(archivo, data_only=True)
    filas_raw = list(_hoja_de_datos(wb).iter_rows(values_only=True))
    if not filas_raw:
        return []
    encabezados = [str(c).strip().lower() if c else '' for c in filas_raw[0]]
    filas = []
    for fila in filas_raw[1:]:
        if all(c is None or str(c).strip() == '' for c in fila):
            continue
        registro = {encabezados[i]: (str(fila[i]).strip() if fila[i] is not None else '') for i in range(len(encabezados))}
        filas.append(registro)
    return filas


# Nombre interno -> encabezados aceptados, en orden de preferencia. Los alias
# existen porque el administrador rara vez usa la plantilla tal cual: renombra
# "correo_electronico" a "correo", "cedula_identidad" a "cedula", etc.
ALIAS_COLUMNAS = {
    'unidad': ('unidad',),
    'rol_ocupacion': ('rol_ocupacion', 'rol'),
    'nombre_completo': ('nombre_completo', 'nombre'),
    'cedula_identidad': ('cedula_identidad', 'cedula'),
    'domicilio': ('domicilio',),
    'correo_electronico': ('correo_electronico', 'correo', 'email'),
    'telefono': ('telefono',),
    'permanencia': ('permanencia',),
    'vinculo_copropietario': ('vinculo_copropietario', 'vinculo'),
    'condicion_especial': ('condicion_especial', 'condicion'),
}

# Se guardan en mayusculas porque son codigos, no texto libre.
CAMPOS_EN_MAYUSCULAS = ('rol_ocupacion', 'permanencia', 'vinculo_copropietario', 'condicion_especial')


def _campos_de_fila(fila):
    """Extrae los campos de una fila cruda, resolviendo alias de encabezado."""
    campos = {}
    for campo, encabezados in ALIAS_COLUMNAS.items():
        valor = ''
        for encabezado in encabezados:
            valor = (fila.get(encabezado) or '').strip()
            if valor:
                break
        campos[campo] = valor.upper() if campo in CAMPOS_EN_MAYUSCULAS else valor

    # Columnas opcionales: vacias significan permanente y sin vinculo declarado.
    campos['permanencia'] = campos['permanencia'] or Permanencia.PERMANENTE
    return campos


def _errores_de_fila(campos):
    """
    Devuelve todos los problemas de una fila, no solo el primero: el
    administrador corrige el archivo una vez y no de a un error por vez.
    """
    errores = []

    for campo, mensaje in (
        ('unidad', 'unidad vacia'),
        ('nombre_completo', 'nombre_completo vacio'),
        ('cedula_identidad', 'cedula_identidad vacia'),
        ('domicilio', 'domicilio vacio'),
    ):
        if not campos[campo]:
            errores.append(mensaje)

    if campos['rol_ocupacion'] not in ROLES_VALIDOS:
        errores.append(
            f"rol_ocupacion invalido: '{campos['rol_ocupacion']}' "
            f"(use {', '.join(sorted(ROLES_VALIDOS))})"
        )
    if campos['permanencia'] not in PERMANENCIAS_VALIDAS:
        errores.append(
            f"permanencia invalida: '{campos['permanencia']}' (use PERMANENTE o TRANSITORIO)"
        )
    if campos['vinculo_copropietario'] and campos['vinculo_copropietario'] not in VINCULOS_VALIDOS:
        errores.append(
            f"vinculo_copropietario invalido: '{campos['vinculo_copropietario']}' "
            f"(use CONYUGE, CONVIVIENTE_CIVIL o FAMILIAR)"
        )
    if campos['condicion_especial'] and campos['condicion_especial'] not in CONDICIONES_VALIDAS:
        errores.append(
            f"condicion_especial invalida: '{campos['condicion_especial']}' "
            f"(use FALLECIDO, DISCAPACIDAD o REQUIERE_APOYO)"
        )

    correo = campos['correo_electronico']
    if not correo:
        errores.append('correo_electronico vacio')
    else:
        try:
            validate_email(correo)
        except ValidationError:
            errores.append(f"correo_electronico invalido: '{correo}'")

    return errores


def _guardar_persona(condominio, campos):
    """Crea o actualiza a la persona. La unidad se crea sola si no existia."""
    unidad, _ = Unidad.objects.get_or_create(
        condominio=condominio, identificador=campos['unidad'],
    )
    Persona.objects.update_or_create(
        condominio=condominio,
        unidad=unidad,
        cedula_identidad=campos['cedula_identidad'],
        defaults={
            'rol_ocupacion': campos['rol_ocupacion'],
            'permanencia': campos['permanencia'],
            'vinculo_copropietario': campos['vinculo_copropietario'],
            'condicion_especial': campos['condicion_especial'],
            'nombre_completo': campos['nombre_completo'],
            'domicilio': campos['domicilio'],
            'correo_electronico': campos['correo_electronico'],
            'telefono': campos['telefono'],
            'activo': True,
        },
    )


def importar_registro_copropietarios(condominio, archivo):
    """
    Procesa el archivo subido, validando cada fila de forma independiente: una
    fila mala nunca impide importar las buenas.

    Devuelve (filas_totales, filas_ok, filas_error, detalle_errores). El numero
    de fila del detalle es el que ve el administrador en Excel, contando el
    encabezado, para que pueda ir directo a corregirla.
    """
    filas = _leer_filas(archivo)
    filas_ok = 0
    detalle_errores = []

    for numero, fila in enumerate(filas, start=2):
        campos = _campos_de_fila(fila)
        errores = _errores_de_fila(campos)
        if errores:
            detalle_errores.append({'fila': numero, 'errores': errores})
            continue
        _guardar_persona(condominio, campos)
        filas_ok += 1

    return len(filas), filas_ok, len(detalle_errores), detalle_errores
