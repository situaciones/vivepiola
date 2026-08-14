"""Tests de la carga del registro de copropietarios (Excel/CSV)."""

import tempfile
from decimal import Decimal

from django.core import mail
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import override_settings
from django.utils import timezone
from rest_framework.test import APITestCase

from accounts.models import Rol, Usuario
from condominios.models import (
    Condominio, Permanencia, Persona, RolOcupacion, Unidad, VinculoCopropietario,
)
from gastos_comunes.utils import exportar_multas_firmes
from multas.models import EstadoMulta, Multa, Ticket, TipoActo
from reglamentos.models import EstadoInfraccion, InfraccionCatalogo

MEDIA_TEMP = tempfile.mkdtemp(prefix='debido_test_media_')

CSV_VALIDO = (
    'unidad,rol_ocupacion,nombre_completo,cedula_identidad,domicilio,correo_electronico,telefono\n'
    'Depto 301,PROPIETARIO,Ana Rojas,10.111.222-3,Depto 301,ana@test.cl,+56911111111\n'
    'Depto 301,ARRENDATARIO,Luis Diaz,12.333.444-5,Depto 301,luis@test.cl,\n'
    'Depto 402,OCUPANTE,Mario Vera,14.555.666-7,Depto 402,mario@test.cl,\n'
)

CSV_CON_ERRORES = (
    'unidad,rol_ocupacion,nombre_completo,cedula_identidad,domicilio,correo_electronico,telefono\n'
    'Depto 501,PROPIETARIO,Carla Munoz,15.111.222-3,Depto 501,carla@test.cl,\n'
    'Depto 502,DUENO,Rol Invalido,16.111.222-3,Depto 502,rol@test.cl,\n'
    'Depto 503,OCUPANTE,Sin Correo,17.111.222-3,Depto 503,correo-invalido,\n'
)


@override_settings(MEDIA_ROOT=MEDIA_TEMP)
class ImportacionRegistroTestCase(APITestCase):
    @classmethod
    def setUpTestData(cls):
        cls.condominio = Condominio.objects.create(nombre='Condominio Import')
        cls.administrador = Usuario.objects.create_user(
            username='admin_import', password='x', rol=Rol.ADMINISTRADOR, condominio=cls.condominio,
        )
        cls.residente = Usuario.objects.create_user(
            username='residente_import', password='x', rol=Rol.RESIDENTE, condominio=cls.condominio,
        )

    def importar(self, contenido, usuario=None):
        self.client.force_authenticate(usuario or self.administrador)
        archivo = SimpleUploadedFile('registro.csv', contenido.encode('utf-8'), content_type='text/csv')
        return self.client.post('/api/registro/importar/', {'archivo': archivo}, format='multipart')

    def test_importacion_valida_crea_unidades_y_personas(self):
        respuesta = self.importar(CSV_VALIDO)
        self.assertEqual(respuesta.status_code, 201, respuesta.data)
        self.assertEqual(respuesta.data['filas_ok'], 3)
        self.assertEqual(respuesta.data['filas_error'], 0)
        self.assertEqual(Unidad.objects.filter(condominio=self.condominio).count(), 2)
        self.assertEqual(Persona.objects.filter(condominio=self.condominio).count(), 3)

    def test_importacion_reporta_filas_invalidas_sin_abortar(self):
        respuesta = self.importar(CSV_CON_ERRORES)
        self.assertEqual(respuesta.status_code, 201)
        self.assertEqual(respuesta.data['filas_ok'], 1)
        self.assertEqual(respuesta.data['filas_error'], 2)
        self.assertEqual(respuesta.data['estado'], 'CON_ERRORES')
        errores = {e['fila'] for e in respuesta.data['detalle_errores']}
        self.assertEqual(errores, {3, 4})

    def test_reimportar_actualiza_sin_duplicar(self):
        self.importar(CSV_VALIDO)
        respuesta = self.importar(CSV_VALIDO)
        self.assertEqual(respuesta.status_code, 201)
        self.assertEqual(Persona.objects.filter(condominio=self.condominio).count(), 3)

    def test_solo_administrador_importa(self):
        respuesta = self.importar(CSV_VALIDO, usuario=self.residente)
        self.assertEqual(respuesta.status_code, 403)

    # -- Titulos de la Ley 21.442 y permanencia ------------------------

    def test_importa_usufructuario_y_comodatario(self):
        contenido = (
            'unidad,rol_ocupacion,nombre_completo,cedula_identidad,domicilio,correo_electronico,telefono\n'
            'Depto 601,USUFRUCTUARIO,Carla Usufructo,18.111.222-3,Depto 601,carla.u@test.cl,\n'
            'Depto 602,COMODATARIO,Dario Comodato,19.111.222-3,Depto 602,dario.c@test.cl,\n'
        )
        respuesta = self.importar(contenido)
        self.assertEqual(respuesta.data['filas_ok'], 2, respuesta.data['detalle_errores'])
        self.assertEqual(
            Persona.objects.get(cedula_identidad='18.111.222-3').rol_ocupacion, RolOcupacion.USUFRUCTUARIO,
        )

    def test_la_permanencia_es_independiente_del_titulo(self):
        contenido = (
            'unidad,rol_ocupacion,nombre_completo,cedula_identidad,domicilio,correo_electronico,telefono,permanencia\n'
            'Depto 701,ARRENDATARIO,Fija Larga,21.111.222-3,Depto 701,fija@test.cl,,PERMANENTE\n'
            'Depto 702,ARRENDATARIO,Turista Corta,22.111.222-3,Depto 702,turista@test.cl,,TRANSITORIO\n'
        )
        self.importar(contenido)
        self.assertEqual(
            Persona.objects.get(cedula_identidad='21.111.222-3').permanencia, Permanencia.PERMANENTE,
        )
        self.assertEqual(
            Persona.objects.get(cedula_identidad='22.111.222-3').permanencia, Permanencia.TRANSITORIO,
        )

    def test_permanencia_por_defecto_si_la_columna_no_viene(self):
        self.importar(CSV_VALIDO)  # el CSV clasico no trae la columna
        self.assertEqual(
            Persona.objects.get(cedula_identidad='10.111.222-3').permanencia, Permanencia.PERMANENTE,
        )

    def test_permanencia_invalida_rechaza_la_fila(self):
        contenido = (
            'unidad,rol_ocupacion,nombre_completo,cedula_identidad,domicilio,correo_electronico,telefono,permanencia\n'
            'Depto 801,PROPIETARIO,Mala Permanencia,23.111.222-3,Depto 801,mala@test.cl,,A VECES\n'
        )
        respuesta = self.importar(contenido)
        self.assertEqual(respuesta.data['filas_error'], 1)
        self.assertIn('permanencia invalida', respuesta.data['detalle_errores'][0]['errores'][0])

    def test_acepta_los_encabezados_que_el_administrador_suele_renombrar(self):
        """Casi nadie usa la plantilla tal cual: renombra correo, cedula, rol."""
        contenido = (
            'unidad,rol,nombre,cedula,domicilio,correo,telefono\n'
            'Depto 111,PROPIETARIO,Alias Renombrado,25.111.222-3,Depto 111,alias@test.cl,\n'
        )
        respuesta = self.importar(contenido)
        self.assertEqual(respuesta.data['filas_ok'], 1, respuesta.data['detalle_errores'])
        persona = Persona.objects.get(cedula_identidad='25.111.222-3')
        self.assertEqual(persona.correo_electronico, 'alias@test.cl')
        self.assertEqual(persona.rol_ocupacion, RolOcupacion.PROPIETARIO)

    def test_una_fila_reporta_todos_sus_errores_de_una_vez(self):
        """Corregir el archivo una vez, no de a un error por vez."""
        contenido = (
            'unidad,rol_ocupacion,nombre_completo,cedula_identidad,domicilio,correo_electronico,telefono\n'
            ',NOEXISTE,,,,correo-malo,\n'
        )
        respuesta = self.importar(contenido)
        errores = respuesta.data['detalle_errores'][0]['errores']
        self.assertGreaterEqual(len(errores), 6, errores)
        texto = ' | '.join(errores)
        for esperado in ('unidad vacia', 'rol_ocupacion invalido', 'nombre_completo vacio',
                         'cedula_identidad vacia', 'domicilio vacio', 'correo_electronico invalido'):
            self.assertIn(esperado, texto)

    def test_registra_el_vinculo_con_el_copropietario(self):
        contenido = (
            'unidad,rol_ocupacion,nombre_completo,cedula_identidad,domicilio,correo_electronico,telefono,permanencia,vinculo_copropietario\n'
            'Depto 901,OCUPANTE,Conyuge Del Dueno,24.111.222-3,Depto 901,conyuge@test.cl,,PERMANENTE,CONYUGE\n'
        )
        respuesta = self.importar(contenido)
        self.assertEqual(respuesta.data['filas_ok'], 1, respuesta.data['detalle_errores'])
        self.assertEqual(
            Persona.objects.get(cedula_identidad='24.111.222-3').vinculo_copropietario, 'CONYUGE',
        )


@override_settings(MEDIA_ROOT=MEDIA_TEMP)
class ObligadoAlPagoTestCase(APITestCase):
    """
    La Ley 21.442 hace al copropietario obligado principal al pago: si multan
    al arrendatario, el cargo del gasto comun igual se emite a nombre del dueño.
    """

    @classmethod
    def setUpTestData(cls):
        cls.condominio = Condominio.objects.create(nombre='Condominio Cobro')
        cls.unidad = Unidad.objects.create(condominio=cls.condominio, identificador='Depto 500')
        cls.dueno = Persona.objects.create(
            condominio=cls.condominio, unidad=cls.unidad, rol_ocupacion=RolOcupacion.PROPIETARIO,
            nombre_completo='Elena Propietaria', cedula_identidad='10.000.000-1',
            domicilio='Depto 500', correo_electronico='elena@test.local',
        )
        cls.arrendatario = Persona.objects.create(
            condominio=cls.condominio, unidad=cls.unidad, rol_ocupacion=RolOcupacion.ARRENDATARIO,
            nombre_completo='Fabian Arrendatario', cedula_identidad='20.000.000-2',
            domicilio='Depto 500', correo_electronico='fabian@test.local',
        )
        cls.admin = Usuario.objects.create_user(
            username='admin_cobro', password='x', rol=Rol.ADMINISTRADOR, condominio=cls.condominio,
        )
        cls.infraccion = InfraccionCatalogo.objects.create(
            condominio=cls.condominio, codigo='RUIDO-01', descripcion='Ruidos molestos',
            articulo_referencia='Art. 15', monto=Decimal('3.00'), estado=EstadoInfraccion.ACTIVA,
        )

    def _multa_firme(self, infractor):
        ticket = Ticket.objects.create(
            condominio=self.condominio, unidad=self.unidad, persona_reportada=infractor,
            descripcion='Hecho', fecha_hecho=timezone.now(),
        )
        return Multa.objects.create(
            condominio=self.condominio, ticket=ticket, unidad=self.unidad,
            persona_infractor=infractor, infraccion=self.infraccion,
            monto=Decimal('3.00'), estado=EstadoMulta.FIRME,
        )

    def test_la_unidad_conoce_a_su_propietario(self):
        self.assertEqual(self.unidad.propietario, self.dueno)

    def test_el_cobro_va_al_dueno_aunque_multen_al_arrendatario(self):
        self._multa_firme(self.arrendatario)
        lote = exportar_multas_firmes(self.condominio, '2026-07', self.admin)

        contenido = lote.archivo_csv.read().decode('utf-8-sig')
        encabezado, fila = contenido.strip().splitlines()[:2]

        self.assertIn('obligado_al_pago', encabezado)
        self.assertIn('Elena Propietaria', fila, 'el cargo debe emitirse al dueño de la unidad')
        self.assertIn('10.000.000-1', fila)
        # El infractor se conserva para que el cargo sea explicable.
        self.assertIn('Fabian Arrendatario', fila)

    def test_sin_propietario_registrado_responde_el_infractor(self):
        self.dueno.delete()
        self._multa_firme(self.arrendatario)
        lote = exportar_multas_firmes(self.condominio, '2026-08', self.admin)
        self.assertIn('Fabian Arrendatario', lote.archivo_csv.read().decode('utf-8-sig'))


class PlantillaExcelTestCase(APITestCase):
    """
    La plantilla que descarga el administrador es la unica instruccion que
    lee antes de llenar el registro: si no explica lo que provocan las
    columnas, quedan sin llenar y las reglas que dependen de ellas no corren.
    """

    def test_la_plantilla_trae_las_columnas_que_el_importador_entiende(self):
        import openpyxl

        from condominios.utils import COLUMNAS_PLANTILLA, generar_plantilla_excel

        ws = openpyxl.load_workbook(generar_plantilla_excel())['Registro Copropietarios']
        encabezados = list(next(ws.iter_rows(max_row=1, values_only=True)))
        self.assertEqual(encabezados, COLUMNAS_PLANTILLA)
        self.assertEqual(ws.max_row, 1, 'sin filas de ejemplo: se importarian como personas reales')

    def test_las_instrucciones_explican_a_quien_le_llega_la_notificacion(self):
        import openpyxl

        from condominios.utils import generar_plantilla_excel

        ws = openpyxl.load_workbook(generar_plantilla_excel())['Instrucciones']
        texto = ' '.join(
            str(c) for fila in ws.iter_rows(values_only=True) for c in fila if c
        )
        self.assertIn('A QUIEN LE LLEGA LA NOTIFICACION', texto)
        self.assertIn('TRANSITORIO', texto)
        self.assertIn('copie al propietario', texto)

    def test_los_titulos_van_en_negrita_aunque_se_agreguen_renglones(self):
        import openpyxl

        from condominios.utils import generar_plantilla_excel

        ws = openpyxl.load_workbook(generar_plantilla_excel())['Instrucciones']
        negritas = {
            str(fila[0].value) for fila in ws.iter_rows(min_col=1, max_col=1)
            if fila[0].value and fila[0].font.bold
        }
        self.assertEqual(
            negritas,
            {'COMO LLENAR ESTA PLANILLA', 'COLUMNA', 'QUIEN PAGA',
             'A QUIEN LE LLEGA LA NOTIFICACION', 'ERRORES QUE RECHAZAN LA FILA'},
        )


@override_settings(MEDIA_ROOT=MEDIA_TEMP)
class CopiaAlCopropietarioTestCase(APITestCase):
    """
    A quien le llega la notificacion legal.

    Se copia al copropietario cuando el infractor podria no estar para ejercer
    su defensa (permanencia transitoria) o cuando ocupa la unidad por su
    vinculo con el dueño. En el resto de los casos basta con el infractor: la
    copia sistematica seria exponer sin necesidad.
    """

    @classmethod
    def setUpTestData(cls):
        cls.condominio = Condominio.objects.create(nombre='Condominio Copia')
        cls.unidad = Unidad.objects.create(condominio=cls.condominio, identificador='Depto 700')
        cls.dueno = Persona.objects.create(
            condominio=cls.condominio, unidad=cls.unidad, rol_ocupacion=RolOcupacion.PROPIETARIO,
            nombre_completo='Rosa Dueña', cedula_identidad='9.000.000-1',
            domicilio='Depto 700', correo_electronico='rosa@test.local',
        )
        cls.admin = Usuario.objects.create_user(
            username='admin_copia', password='x', rol=Rol.ADMINISTRADOR, condominio=cls.condominio,
        )
        cls.infraccion = InfraccionCatalogo.objects.create(
            condominio=cls.condominio, codigo='PISCINA-01', descripcion='Uso de piscina fuera de horario',
            articulo_referencia='Art. 22', monto=Decimal('2.00'), estado=EstadoInfraccion.ACTIVA,
        )

    def _notificar(self, infractor):
        from multas.services import notificar_multa

        ticket = Ticket.objects.create(
            condominio=self.condominio, unidad=self.unidad, persona_reportada=infractor,
            descripcion='Uso de piscina a las 3 AM', fecha_hecho=timezone.now(),
        )
        multa = Multa.objects.create(
            condominio=self.condominio, ticket=ticket, unidad=self.unidad,
            persona_infractor=infractor, infraccion=self.infraccion,
            monto=Decimal('2.00'), estado=EstadoMulta.APROBADA,
        )
        mail.outbox = []
        notificar_multa(multa, self.admin)
        return multa, mail.outbox[0]

    def _persona(self, **kwargs):
        datos = dict(
            condominio=self.condominio, unidad=self.unidad, rol_ocupacion=RolOcupacion.ARRENDATARIO,
            nombre_completo='Ocupante', cedula_identidad='9.111.111-1',
            domicilio='Depto 700', correo_electronico='ocupante@test.local',
        )
        datos.update(kwargs)
        return Persona.objects.create(**datos)

    def test_al_transitorio_se_le_copia_al_dueno(self):
        """Puede haberse ido antes de que corra el plazo de descargo."""
        turista = self._persona(
            nombre_completo='Tomas Turista', cedula_identidad='9.222.222-2',
            correo_electronico='tomas@test.local', permanencia=Permanencia.TRANSITORIO,
        )
        _, correo = self._notificar(turista)

        self.assertEqual(correo.to, ['tomas@test.local'])
        self.assertEqual(correo.cc, ['rosa@test.local'])
        self.assertIn('Rosa Dueña', correo.body, 'el correo debe decir a quien se copio y por que')

    def test_al_permanente_sin_vinculo_no_se_copia_a_nadie(self):
        arrendatario = self._persona(
            nombre_completo='Pablo Permanente', cedula_identidad='9.333.333-3',
            correo_electronico='pablo@test.local', permanencia=Permanencia.PERMANENTE,
        )
        _, correo = self._notificar(arrendatario)

        self.assertEqual(correo.to, ['pablo@test.local'])
        self.assertEqual(correo.cc, [], 'copiar al dueño siempre seria exponer sin necesidad')

    def test_a_quien_ocupa_por_vinculo_con_el_dueno_se_le_copia(self):
        conyuge = self._persona(
            rol_ocupacion=RolOcupacion.OCUPANTE, nombre_completo='Carmen Conyuge',
            cedula_identidad='9.444.444-4', correo_electronico='carmen@test.local',
            vinculo_copropietario=VinculoCopropietario.CONYUGE,
        )
        _, correo = self._notificar(conyuge)
        self.assertEqual(correo.cc, ['rosa@test.local'])

    def test_al_dueno_no_se_le_copia_su_propia_multa(self):
        self.dueno.permanencia = Permanencia.TRANSITORIO
        self.dueno.save(update_fields=['permanencia'])
        _, correo = self._notificar(self.dueno)

        self.assertEqual(correo.to, ['rosa@test.local'])
        self.assertEqual(correo.cc, [], 'nadie se copia a si mismo')

    def test_sin_correo_del_dueno_la_notificacion_igual_sale(self):
        """La copia es un refuerzo: nunca puede impedir el canal legal principal."""
        self.dueno.correo_electronico = ''
        self.dueno.save(update_fields=['correo_electronico'])
        turista = self._persona(
            nombre_completo='Tania Transitoria', cedula_identidad='9.555.555-5',
            correo_electronico='tania@test.local', permanencia=Permanencia.TRANSITORIO,
        )
        _, correo = self._notificar(turista)

        self.assertEqual(correo.to, ['tania@test.local'])
        self.assertEqual(correo.cc, [])

    def test_el_acta_sella_a_quien_se_copio(self):
        """Si mañana se discute si el copropietario fue notificado, el acta lo prueba."""
        turista = self._persona(
            nombre_completo='Teo Transitorio', cedula_identidad='9.666.666-6',
            correo_electronico='teo@test.local', permanencia=Permanencia.TRANSITORIO,
        )
        multa, _ = self._notificar(turista)

        acta = multa.actas_selladas.filter(tipo_acto=TipoActo.NOTIFICACION).latest('id')
        self.assertEqual(acta.manifiesto['extra']['copias_copropietario'], ['rosa@test.local'])
